from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SCHOOL_NAME = "TRƯỜNG ĐẠI HỌC CÔNG NGHỆ KỸ THUẬT TP.HCM"
DEPARTMENT_NAME = "PHÒNG ĐÀO TẠO"
DEFAULT_CREDITS = 3
NO_DATA_TEXT = "Không có dữ liệu"
ABSENT_TEXT = "Vắng mặt"


def export_session_attendance_excel(file_path, export_data):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Diem danh"
    _write_attendance_sheet(worksheet, export_data)
    workbook.save(file_path)


def export_all_sessions_attendance_excel(file_path, export_data_items):
    workbook = Workbook()
    workbook.remove(workbook.active)

    for index, export_data in enumerate(export_data_items, start=1):
        session = export_data.get("session", {})
        course_code = session.get("course_code") or f"Buoi {index}"
        worksheet = workbook.create_sheet(title=_safe_sheet_title(str(course_code), index))
        _write_attendance_sheet(worksheet, export_data)

    workbook.save(file_path)


def _write_attendance_sheet(worksheet, export_data):
    session = export_data.get("session", {})
    students = export_data.get("students", [])
    headers = [
        "TT",
        "Mã SV",
        "Họ và tên lót",
        "Tên",
        "Ngày sinh",
        "Tên lớp",
        "Điểm danh",
        "Trạng thái học tập",
    ]
    last_col = len(headers)
    last_col_letter = get_column_letter(last_col)

    worksheet.sheet_view.showGridLines = False
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.bottom = 0.5

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    worksheet["A1"] = SCHOOL_NAME
    worksheet["A1"].font = Font(name="Times New Roman", size=13, bold=True)
    worksheet["A1"].alignment = Alignment(horizontal="center")

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    worksheet["A2"] = DEPARTMENT_NAME
    worksheet["A2"].font = Font(name="Times New Roman", size=13, bold=True)
    worksheet["A2"].alignment = Alignment(horizontal="center")

    worksheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
    worksheet["A4"] = "DANH SÁCH ĐIỂM DANH"
    worksheet["A4"].font = Font(name="Times New Roman", size=18, bold=True)
    worksheet["A4"].alignment = Alignment(horizontal="center")

    worksheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=last_col)
    worksheet["A5"] = _semester_line(session)
    worksheet["A5"].font = Font(name="Times New Roman", size=13)
    worksheet["A5"].alignment = Alignment(horizontal="center")

    worksheet.merge_cells(start_row=6, start_column=1, end_row=6, end_column=6)
    worksheet["A6"] = _course_line(session)
    worksheet["A6"].font = Font(name="Times New Roman", size=13, bold=True)

    worksheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=5)
    worksheet["A7"] = _teacher_line(session)
    worksheet["A7"].font = Font(name="Times New Roman", size=13, bold=True)

    worksheet.merge_cells(start_row=7, start_column=7, end_row=7, end_column=last_col)
    worksheet["G7"] = f"Số tín chỉ: {session.get('credits') or DEFAULT_CREDITS}"
    worksheet["G7"].font = Font(name="Times New Roman", size=13, bold=True)
    worksheet["G7"].alignment = Alignment(horizontal="right")

    header_row = 9
    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=col_index, value=header)
        cell.font = Font(name="Times New Roman", size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")

    for row_index, student in enumerate(students, start=1):
        excel_row = header_row + row_index
        family_name, given_name = _split_vietnamese_name(student.get("full_name") or "")
        values = [
            row_index,
            student.get("student_id") or "",
            family_name,
            given_name,
            "",
            student.get("class_name") or session.get("class_name") or "",
            _attendance_text(student.get("attendance_status")),
            student.get("latest_behavior") or NO_DATA_TEXT,
        ]
        for col_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=excel_row, column=col_index, value=value)
            cell.font = Font(name="Times New Roman", size=12)
            cell.alignment = Alignment(
                horizontal="center" if col_index in (1, 2, 5, 7) else "left",
                vertical="center",
                wrap_text=True,
            )

    table_last_row = max(header_row, header_row + len(students))
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in worksheet.iter_rows(min_row=header_row, max_row=table_last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = border

    widths = [6, 14, 26, 14, 13, 20, 16, 24]
    for col_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(col_index)].width = width

    for row_number in range(1, table_last_row + 1):
        worksheet.row_dimensions[row_number].height = 22
    worksheet.row_dimensions[4].height = 28
    worksheet.freeze_panes = "A10"

    signature_row = table_last_row + 3
    worksheet.merge_cells(start_row=signature_row, start_column=6, end_row=signature_row, end_column=last_col)
    signature_cell = worksheet.cell(row=signature_row, column=6, value=_signature_date(session.get("lecture_date")))
    signature_cell.font = Font(name="Times New Roman", size=12)
    signature_cell.alignment = Alignment(horizontal="center")

    worksheet.auto_filter.ref = f"A{header_row}:{last_col_letter}{table_last_row}"


def _semester_line(session):
    academic_year = session.get("academic_year") or _academic_year_from_date(session.get("lecture_date"))
    semester = session.get("semester") or ""
    if semester:
        return f"Học kỳ: {semester} - Năm học: {academic_year}"
    return f"Năm học: {academic_year}" if academic_year else "Năm học:"


def _course_line(session):
    course_name = session.get("course_name") or ""
    course_code = session.get("course_code") or ""
    class_name = session.get("class_name") or session.get("class_code") or ""
    course_part = f"{course_name} ({course_code})" if course_code else course_name
    group_part = f" - Nhóm {class_name}" if class_name else ""
    return f"Môn học/Nhóm: {course_part}{group_part}"


def _teacher_line(session):
    teacher_name = session.get("teacher_name") or "Chưa có dữ liệu"
    teacher_id = session.get("teacher_id")
    teacher_code = f"({teacher_id})" if teacher_id else ""
    return f"CBGD: {teacher_name}{teacher_code}"


def _split_vietnamese_name(full_name):
    parts = [part for part in full_name.strip().split() if part]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return "", parts[0]
    return " ".join(parts[:-1]), parts[-1]


def _attendance_text(status):
    return status or ABSENT_TEXT


def _academic_year_from_date(date_text):
    try:
        year = datetime.strptime(str(date_text), "%Y-%m-%d").year
        return f"{year}-{year + 1}"
    except (TypeError, ValueError):
        return ""


def _signature_date(date_text):
    try:
        date_value = datetime.strptime(str(date_text), "%Y-%m-%d")
    except (TypeError, ValueError):
        date_value = datetime.now()
    return f"Ngày {date_value.day} tháng {date_value.month} năm {date_value.year}"


def _safe_sheet_title(title, index):
    cleaned = "".join("_" if char in r'[]:*?/\\' else char for char in title).strip()
    cleaned = cleaned[:25] or "Diem danh"
    return f"{index:02d}_{cleaned}"[:31]
